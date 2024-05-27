import pandas as pd
import openpyxl

class GetDataFromExcel():
    excel_data_sheet = openpyxl.load_workbook("Path_To_Excel")['Sheet1']
    length = excel_data_sheet.max_row - 1
    
    def __init__(self, restricted_brands: list) -> None:
        self.restricted_brands = restricted_brands
        self.table_data = {
                "Name" : [],
                "P_Id" : [],
                "Brand" : [],
                "Our_Price" : [],
                "Lowest_Price" : [],
                "Link" : [],
                "Status": []
                }
        self.extractData()

    def restrictedBrands(self, brand: str) -> bool:
        if brand in self.restricted_brands:
            return True

    def extractData(self):
        for row in range(2, self.excel_data_sheet.max_row+1):
            try:
                brand = self.excel_data_sheet.cell(row=row, column=6).value
                if self.restricted_brands(brand):
                    continue
                else:
                    self.table_data["Link"].append(brand)
            except Exception as e:
                print(f"Cannot get Brand at row: {row} due to error:")
                print(e)
                continue
            try:
                self.table_data["Name"].append(self.excel_data_sheet.cell(row=row, column=1).value)
            except Exception as e:
                self.table_data["Name"].append("")
                print(f"Cannot get Name at row: {row} due to error:")
                print(e)
            try:
                self.table_data["P_Id"].append(self.excel_data_sheet.cell(row=row, column=2).value)
            except Exception as e:
                self.table_data["P_Id"].append("")
                print(f"Cannot get P_Id at row: {row} due to error:")
                print(e)
            try:
                self.table_data["Our_Price"].append(self.excel_data_sheet.cell(row=row, column=6).value)
            except Exception as e:
                self.table_data["Our_Price"].append("")
                print(f"Cannot get Our_Price at row: {row} due to error:")
                print(e)
            try:
                self.table_data["Lowest_Price"].append(self.excel_data_sheet.cell(row=row, column=6).value)
            except Exception as e:
                self.table_data["Lowest_Price"].append("")
                print(f"Cannot get Lowest_Price at row: {row} due to error:")
                print(e)
            try:
                self.table_data["link"].append(self.excel_data_sheet.cell(row=row, column=6).value)
            except Exception as e:
                self.table_data["link"].append("")
                self.table_data["Status"].append("Invalid Link")
                print(f"Cannot get Link at row: {row} due to error:")
                print(e)
                continue
            try:
                self.table_data["Status"].append(self.excel_data_sheet.cell(row=row, column=6).value)
            except Exception as e:
                self.table_data["Status"].append("")
                print(f"Cannot get Status at row: {row} due to error:")
                print(e)