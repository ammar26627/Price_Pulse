import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pytz import timezone 
from datetime import datetime

class PushDataToExcel():
    PATH = "Path of file/"
    link_sheet_excel = load_workbook(PATH)['Sheet1']

    def __init__(self, table_data: dict) -> None:
        self.table_data = table_data
        self.current_path = self.PATH + datetime.now(timezone("Asia/Kolkata")).strftime('%Y-%m-%d_%H:%M:%S')
        self.formatExcel()
    
    def pushData(self):
        df = pd.DataFrame(self.table_data)
        df.to_excel(self.current_path, index=False)
    
    def fillColor(row, color):
        for cell in row:
                cell.fill = color

    def formatExcel(self):
        green_fill = PatternFill(start_color='CCFEFF', end_color='CCFEFF', fill_type='solid')
        red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF6E', end_color='FFFF6E', fill_type='solid')

        for row in self.link_sheet_excel.iter_rows(min_row=2, max_row=self.link_sheet_excel.max_row, min_col=1, max_col=self.link_sheet_excel.max_column):
            if row[5].value != "":
                if row[5].value >= row[4]:
                    self.fillColor(row, green_fill)
                else:
                    self.fillColor(row, red_fill)
            else:
                self.fillColor(row, yellow_fill)

        self.link_sheet_excel.save(self.current_path)