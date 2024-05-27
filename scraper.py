import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
import requests

def test():
    print("Test")

def scrape(excel, given_brands):
    otherSellerPrice = []
    orginalPrice = []
    productLinks = []
    productId = []
    errorList = []
    global loadPercent
    global inter
    global df
    inter = False
    loadPercent = 0
    load = 0
    link_sheet_workbook = openpyxl.load_workbook('./Data/Excel/{}.xlsx'.format(excel))
    link_sheet_excel = link_sheet_workbook['Sheet1']
    length = link_sheet_excel.max_row - 1

    for row in range(2, link_sheet_excel.max_row+1):
        if(inter==True):
                print("break")
                break
        print("Getting...")
        try:
            link = link_sheet_excel.cell(row=row, column=1).hyperlink.target
        except:
            try:
                link = link_sheet_excel.cell(row=row, column=14).value
            except:
                otherSellerPrice.append("")
                try:
                    orginalPrice.append(link_sheet_excel.cell(row=row, column=9).value)
                except:
                    orginalPrice.append("")
                try:
                    productId.append(link_sheet_excel.cell(row=row, column=3).value)
                except:
                    productId.append("")
                try:
                    link_sheet_excel.cell(row=row, column=14).value = ""
                except:
                    print ("Link Not Added")
                errorList.append("Invalid Link")
                print ("Link Not Found")
                productLinks.append("")
                load = load + 1
                continue
        URL = link
        try:
            brand = link_sheet_excel.cell(row=row, column=6).value
        except:
            print("Cannot get brand")
            otherSellerPrice.append("")
            try:
                orginalPrice.append(link_sheet_excel.cell(row=row, column=9).value)
            except:
                orginalPrice.append("")
            try:
                productId.append(link_sheet_excel.cell(row=row, column=3).value)
            except:
                productId.append("")
            errorList.append("Invalid Brand")
            load = load + 1
            continue

        if(brand in given_brands):
            print("Restricted Brand")
            otherSellerPrice.append("")
            productLinks.append(link)
            try:
                orginalPrice.append(link_sheet_excel.cell(row=row, column=9).value)
            except:
                orginalPrice.append("")
            try:
                productId.append(link_sheet_excel.cell(row=row, column=3).value)
            except:
                productId.append("")
            link_sheet_excel.cell(row=row, column=12).value = ""
            errorList.append("Restricted Brand")
            load = load + 1
            continue
        

        try:
            r = requests.get(URL)
        except:
            load = load + 1
            continue
        try:
            soup = BeautifulSoup(r.content, 'html.parser')
            opDiv = soup.find('div', attrs = {'class':'offer_price'})
            offerPriceText = opDiv.find('span', attrs = {'class':'m-w'}).text
            minPrice = float(offerPriceText[1:].replace(',', ''))   
            errorList.append("")
        except:
            minPrice = ""
            errorList.append("Could not find minimum price")
        finally: 
            otherSellerPrice.append(minPrice)
            productLinks.append(link)
            try:
                link_sheet_excel.cell(row=row, column=12).value = minPrice
            except:
                print("Minimum Price Cannot Be Added")
            try:
                link_sheet_excel.cell(row=row, column=14).value = link
            except:
                print("Link Not Added")
            try:
                orginalPrice.append(link_sheet_excel.cell(row=row, column=9).value)
            except:
                orginalPrice.append("")
            try:
                productId.append(link_sheet_excel.cell(row=row, column=3).value)
            except:
                productId.append("")

            load = load + 1
            loadPercent = round(((load/length) * 100), 1)

            if(inter==True):
                print("break")
                break
    link_sheet_workbook.save('../Data/Excel/{}.xlsx'.format(excel))
    data = {'ID': productId, 'Original_Price': orginalPrice, 'Other_Seller_Price': otherSellerPrice, 'Product_Link': productLinks, "Error": errorList}
    df = pd.DataFrame(data)
    return df.reset_index().to_json(orient='records')